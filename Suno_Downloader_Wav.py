# Suno_Downloader_Wav_EXE.py
#
# Features:
# - GUI downloader for Suno playlists/clips
# - MP3 downloads with full tags + artwork
# - WAV downloads via Playwright + Chrome persistent profile (pw_suno_profile next to EXE/script)
#   - WAV saved into the SAME folder as MP3 (Audio) and uses SAME filename base as the MP3
#   - WAV downloads run AFTER everything else (metadata/files/indexes)
#   - WAV tagged to MATCH MP3 tags (ID3v2.4; includes WXXX "ID", TXXX "ID", artwork, lyrics, etc.)
# - WAV also writes RIFF INFO tags for better Windows Explorer visibility (Title/Album/Track/Genre/Comments)
# - Skip duplicates by ID (reads existing MP3/FLAC tags to detect)
# - Index XLSX Title column uses filename base (no extension), thumbnails preserved/added
# - Retag Only mode: fills ONLY missing ID tags on MP3/FLAC, preserves timestamps,
#   updates index titles (does NOT rewrite other tags, does NOT touch Title)
# - In-app scrollable console + run log + separate error log
# - Clear / Pause-Resume / Stop buttons
# - Suno Login (Chrome) button: closes app, launches Chrome via PowerShell login script, then relaunches when Chrome closes
# - Save/Load Session: saves inputs + options into session files stored under pw_suno_profile\sessions
#
# Dependencies:
#   pip install requests pillow openpyxl mutagen playwright
#   py -m playwright install
#
# Build (example):
#   py -m PyInstaller --noconfirm --clean --onefile --noconsole --name "Suno_Playlist_Downloader_wav_support" --collect-all playwright "Suno_Downloader_Wav.py"
#     
#
import os
import re
import sys
import time
import json
import shutil
import queue
import threading
import subprocess
import traceback
import webbrowser
import struct
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse
from typing import Any, Dict, List, Optional, Set, Tuple, TYPE_CHECKING

import tkinter as tk
from tkinter import filedialog, messagebox, font
from tkinter import simpledialog
from tkinter.ttk import Progressbar

import requests
from PIL import Image as PILImage, ImageOps
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage

from mutagen.id3 import (
    ID3, TIT2, TALB, TRCK, TCON, USLT, COMM, APIC,
    WXXX, TXXX, TCOP, ID3NoHeaderError
)
from mutagen.mp3 import MP3

# FLAC retag support
try:
    from mutagen.flac import FLAC
    FLAC_AVAILABLE = True
except Exception:
    FLAC = None  # type: ignore
    FLAC_AVAILABLE = False

# WAV tag support
try:
    from mutagen.wave import WAVE
    WAVE_AVAILABLE = True
except Exception:
    WAVE = None  # type: ignore
    WAVE_AVAILABLE = False

# Playwright WAV automation
try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError
    PLAYWRIGHT_AVAILABLE = True
except Exception:
    sync_playwright = None  # type: ignore
    PWTimeoutError = Exception  # type: ignore
    PLAYWRIGHT_AVAILABLE = False

if TYPE_CHECKING:
    from playwright.sync_api import Page, BrowserContext
else:
    Page = Any  # type: ignore
    BrowserContext = Any  # type: ignore


# ===================== App paths / constants =====================

def _app_dir() -> Path:
    try:
        if getattr(sys, "frozen", False):
            return Path(sys.executable).resolve().parent
    except Exception:
        pass
    return Path(__file__).resolve().parent


APP_DIR = _app_dir()

# Chrome profile next to EXE/script
PERSIST_PROFILE_DIR = APP_DIR / "pw_suno_profile"
BOOTSTRAP_MARK_FILE = PERSIST_PROFILE_DIR / ".bootstrapped"

# Chrome path override stored in profile
CHROME_PATH_FILE = PERSIST_PROFILE_DIR / "chrome_path.txt"

# Session storage
SESSION_DIR = PERSIST_PROFILE_DIR / "sessions"

# PowerShell login script next to EXE/script
SUNO_LOGIN_PS1 = APP_DIR / "suno_login.ps1"

# Logs in profile folder
LOG_DIR = PERSIST_PROFILE_DIR / "logs"
LOG_LATEST = LOG_DIR / "suno_downloader_latest.log"
ERR_LATEST = LOG_DIR / "suno_downloader_errors_latest.log"

# Playwright settings
PW_PER_SONG_TIMEOUT_MS = 180_000
PW_CHROME_ARGS = [
    "--start-minimized",
    "--disable-notifications",
    "--no-first-run",
    "--no-default-browser-check",
]
MORE_MENU_PATH_PREFIX = "M6 14q-.824"

# UI console buffer
UI_LOG_MAX_LINES = 25_000
UI_LOG_QUEUE: "queue.Queue[str]" = queue.Queue()

PROFILE_LOCK_FILES = [
    "SingletonLock",
    "SingletonCookie",
    "SingletonSocket",
    "Lockfile",
]

HEADERS = [
    "Index", "Track #", "Playlist", "Title", "Length", "Artwork", "Genre",
    "Prompt", "Lyrics", "ID", "Created", "Model Version", "Model Name",
    "Type", "Weight", "Creativity", "Image Url", "Audio Url", "Video Url"
]

UUID_RE = re.compile(
    r"[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}",
    re.IGNORECASE,
)

# ===================== UI Style =====================

BG_COLOR         = "#150300"
FG_COLOR         = "#170F1E"
BUTTON_BG        = "#EE56AF"
BUTTON_FG        = "#470000"
ENTRY_BG         = "#30242D"
ENTRY_FG         = "#007A59"
CHECK_TEXT_COLOR = "#FF00FE"
DISABLED_COLOR   = "#510053"
SELECT_COLOR     = "#00FF40"
FONT_FAMILY      = "gabriola"
FONT_SIZE        = 14


# ===================== Logging =====================

class _ConsoleCapture:
    def __init__(self, fhs, tag: str, also=None):
        self.fhs = [fh for fh in (fhs or []) if fh]
        self.tag = tag
        self.also = also
        self._buf = ""

    def write(self, s: str):
        if not s:
            return
        self._buf += s
        while "\n" in self._buf:
            line, self._buf = self._buf.split("\n", 1)
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            out = f"[{ts}] [{self.tag}] {line}\n"

            for fh in self.fhs:
                try:
                    fh.write(out)
                    fh.flush()
                except Exception:
                    pass

            try:
                UI_LOG_QUEUE.put_nowait(out)
            except Exception:
                pass

            if self.also:
                try:
                    self.also.write(line + "\n")
                    self.also.flush()
                except Exception:
                    pass

    def flush(self):
        if self._buf:
            self.write("\n")


_LOG_FH = None
_ERR_FH = None

def setup_profile_logging() -> Path:
    global _LOG_FH, _ERR_FH

    try:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass

    run_log = LOG_DIR / f"suno_downloader_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    err_log = LOG_DIR / f"suno_downloader_errors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    _LOG_FH = open(run_log, "a", encoding="utf-8", errors="replace")
    _ERR_FH = open(err_log, "a", encoding="utf-8", errors="replace")

    try:
        LOG_LATEST.write_text(f"Latest run log: {run_log}\n", encoding="utf-8", errors="replace")
    except Exception:
        pass
    try:
        ERR_LATEST.write_text(f"Latest error log: {err_log}\n", encoding="utf-8", errors="replace")
    except Exception:
        pass

    out_also = getattr(sys, "__stdout__", None)
    err_also = getattr(sys, "__stderr__", None)

    sys.stdout = _ConsoleCapture([_LOG_FH], "OUT", also=out_also)
    sys.stderr = _ConsoleCapture([_LOG_FH, _ERR_FH], "ERR", also=err_also)

    def _log_uncaught(exc_type, exc, tb):
        print("UNCAUGHT EXCEPTION:", file=sys.stderr)
        print("".join(traceback.format_exception(exc_type, exc, tb)), file=sys.stderr)

    sys.excepthook = _log_uncaught

    if hasattr(threading, "excepthook"):
        def _thread_excepthook(args):
            print("THREAD EXCEPTION:", file=sys.stderr)
            print("".join(traceback.format_exception(args.exc_type, args.exc_value, args.exc_traceback)), file=sys.stderr)
        threading.excepthook = _thread_excepthook

    print("===============================================")
    print("Suno Downloader start")
    print(f"Time: {datetime.now()}")
    print(f"App dir: {APP_DIR}")
    print(f"Profile dir: {PERSIST_PROFILE_DIR}")
    print(f"Run log: {run_log}")
    print(f"Error log: {err_log}")
    print(f"Playwright available: {PLAYWRIGHT_AVAILABLE}")
    print(f"FLAC tag support available: {FLAC_AVAILABLE}")
    print(f"WAVE tag support available: {WAVE_AVAILABLE}")
    print("===============================================")
    return run_log


def _bootstrap_message(title: str, text: str):
    print(f"[BOOTSTRAP MESSAGE] {title}: {text}")
    root = tk.Tk()
    root.withdraw()
    try:
        root.attributes("-topmost", True)
    except Exception:
        pass
    try:
        messagebox.showinfo(title, text)
    finally:
        try:
            root.destroy()
        except Exception:
            pass


# ===================== Self-watcher mode (relaunch after login) =====================

def _profile_is_locked(profile_dir: Path) -> bool:
    try:
        for nm in PROFILE_LOCK_FILES:
            if (profile_dir / nm).exists():
                return True
    except Exception:
        return False
    return False


def run_bootstrap_watcher_and_relaunch():
    started_lock = False
    start_wait_deadline = time.time() + 60.0

    while time.time() < start_wait_deadline:
        if _profile_is_locked(PERSIST_PROFILE_DIR):
            started_lock = True
            break
        time.sleep(0.5)

    if started_lock:
        stable_gone = 0
        while True:
            if _profile_is_locked(PERSIST_PROFILE_DIR):
                stable_gone = 0
            else:
                stable_gone += 1
                if stable_gone >= 3:
                    break
            time.sleep(1.0)
    else:
        time.sleep(10.0)

    try:
        creationflags = 0
        try:
            creationflags = subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP
        except Exception:
            creationflags = 0

        if getattr(sys, "frozen", False):
            cmd = [str(Path(sys.executable).resolve())]
        else:
            cmd = [str(Path(sys.executable).resolve()), str(Path(__file__).resolve())]

        subprocess.Popen(
            cmd,
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=creationflags
        )
    except Exception as e:
        print(f"[WATCHER] Relaunch failed: {e}", file=sys.stderr)


def _maybe_run_watcher_mode_from_args():
    if any(a.strip().lower() == "--watch-login" for a in sys.argv[1:]):
        setup_profile_logging()
        print("[WATCHER] Started. Waiting for Chrome profile lock to clear…")
        run_bootstrap_watcher_and_relaunch()
        sys.exit(0)


def _spawn_watcher_detached() -> None:
    try:
        creationflags = 0
        try:
            creationflags = subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP
        except Exception:
            creationflags = 0

        if getattr(sys, "frozen", False):
            watcher_cmd = [str(Path(sys.executable).resolve()), "--watch-login"]
        else:
            watcher_cmd = [str(Path(sys.executable).resolve()), str(Path(__file__).resolve()), "--watch-login"]

        subprocess.Popen(
            watcher_cmd,
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=creationflags
        )
    except Exception as e:
        print(f"[WATCHER] Failed to start watcher: {e}", file=sys.stderr)


def _hard_exit_soon(ms: int = 250) -> None:
    def _exit_now():
        try:
            try:
                sys.stdout.flush()
            except Exception:
                pass
            try:
                sys.stderr.flush()
            except Exception:
                pass
        finally:
            os._exit(0)

    try:
        root = tk._default_root
        if root:
            root.after(ms, _exit_now)
            return
    except Exception:
        pass
    _exit_now()


# ===================== Chrome discovery + prompt =====================

def _read_saved_chrome_path() -> Optional[Path]:
    try:
        if CHROME_PATH_FILE.exists():
            raw = CHROME_PATH_FILE.read_text(encoding="utf-8", errors="ignore").strip().strip('"')
            if raw:
                p = Path(raw)
                if p.exists():
                    return p
    except Exception:
        pass
    return None


def _write_saved_chrome_path(p: Path):
    try:
        PERSIST_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
        CHROME_PATH_FILE.write_text(str(p), encoding="utf-8")
    except Exception:
        pass


def find_chrome_candidates() -> List[Path]:
    candidates: List[Path] = []

    saved = _read_saved_chrome_path()
    if saved:
        candidates.append(saved)

    env_pf = os.environ.get("ProgramFiles", r"C:\Program Files")
    env_pf86 = os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")
    env_local = os.environ.get("LocalAppData", "")

    candidates.extend([
        Path(env_pf) / "Google" / "Chrome" / "Application" / "chrome.exe",
        Path(env_pf86) / "Google" / "Chrome" / "Application" / "chrome.exe",
    ])

    if env_local:
        candidates.extend([
            Path(env_local) / "Google" / "Chrome" / "Application" / "chrome.exe",
            Path(env_local) / "Chromium" / "Application" / "chrome.exe",
        ])

    out: List[Path] = []
    seen: Set[str] = set()
    for p in candidates:
        s = str(p).lower()
        if s in seen:
            continue
        seen.add(s)
        out.append(p)
    return out


def find_chrome_exe() -> Optional[Path]:
    for p in find_chrome_candidates():
        try:
            if p.exists():
                return p
        except Exception:
            pass
    return None


def prompt_for_chrome_exe() -> Optional[Path]:
    root = tk.Tk()
    root.withdraw()
    try:
        root.attributes("-topmost", True)
    except Exception:
        pass

    result: Dict[str, Optional[Path]] = {"path": None}

    win = tk.Toplevel(root)
    win.title("Chrome not found")
    win.configure(bg=BG_COLOR)
    win.resizable(False, False)

    msg = (
        "chrome.exe was not found in the usual locations.\n\n"
        "Choose an option:\n"
        "• Browse to an existing chrome.exe\n"
        "• Download and install Google Chrome"
    )

    lbl = tk.Label(win, text=msg, bg=BG_COLOR, fg=CHECK_TEXT_COLOR, font=(FONT_FAMILY, FONT_SIZE))
    lbl.pack(padx=14, pady=(14, 10))

    btn_row = tk.Frame(win, bg=BG_COLOR)
    btn_row.pack(padx=14, pady=(0, 14), fill=tk.X)

    def _browse():
        path = filedialog.askopenfilename(
            title="Select chrome.exe",
            filetypes=[("chrome.exe", "chrome.exe"), ("Executable", "*.exe"), ("All files", "*.*")]
        )
        if path:
            p = Path(path)
            if p.exists() and p.name.lower() == "chrome.exe":
                result["path"] = p
                _write_saved_chrome_path(p)
                win.destroy()
            else:
                messagebox.showerror("Invalid selection", "Please select a valid chrome.exe")

    def _download():
        try:
            webbrowser.open("https://www.google.com/chrome/")
        except Exception:
            pass
        messagebox.showinfo(
            "Install Chrome",
            "Chrome will open in your browser.\n\n"
            "After installing, relaunch this downloader.\n"
            "If you already installed it, click Browse and point to chrome.exe."
        )

    def _exit():
        win.destroy()

    b1 = tk.Button(btn_row, text="Browse…", command=_browse, bg=BUTTON_BG, fg=BUTTON_FG, font=(FONT_FAMILY, FONT_SIZE))
    b2 = tk.Button(btn_row, text="Download Chrome", command=_download, bg=BUTTON_BG, fg=BUTTON_FG, font=(FONT_FAMILY, FONT_SIZE))
    b3 = tk.Button(btn_row, text="Exit", command=_exit, bg=BUTTON_BG, fg=BUTTON_FG, font=(FONT_FAMILY, FONT_SIZE))

    b1.pack(side=tk.LEFT, padx=(0, 8))
    b2.pack(side=tk.LEFT, padx=(0, 8))
    b3.pack(side=tk.LEFT)

    win.protocol("WM_DELETE_WINDOW", _exit)

    win.update_idletasks()
    w = win.winfo_width()
    h = win.winfo_height()
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    x = (sw // 2) - (w // 2)
    y = (sh // 2) - (h // 2)
    win.geometry(f"{w}x{h}+{x}+{y}")

    win.grab_set()
    root.wait_window(win)

    try:
        root.destroy()
    except Exception:
        pass

    return result["path"]


def ensure_chrome_exe() -> Optional[Path]:
    p = find_chrome_exe()
    if p:
        return p
    return prompt_for_chrome_exe()


# ===================== PowerShell login launcher =====================

def _ensure_suno_login_ps1_exists() -> None:
    if SUNO_LOGIN_PS1.exists():
        return

    content = r"""param(
  [Parameter(Mandatory=$true)][string]$ChromePath,
  [Parameter(Mandatory=$true)][string]$ProfileDir,
  [string]$Url = "https://suno.com/"
)

$ErrorActionPreference = "Stop"

if (-not (Test-Path -LiteralPath $ChromePath)) {
  Write-Host "Chrome not found: $ChromePath"
  exit 2
}

if (-not (Test-Path -LiteralPath $ProfileDir)) {
  New-Item -ItemType Directory -Path $ProfileDir -Force | Out-Null
}

$argList = @(
  "--user-data-dir=$ProfileDir",
  "--no-first-run",
  "--no-default-browser-check",
  "--new-window",
  $Url
)

Start-Process -FilePath $ChromePath -ArgumentList $argList | Out-Null
"""
    try:
        SUNO_LOGIN_PS1.write_text(content, encoding="utf-8", errors="replace")
    except Exception as e:
        print(f"[PS1] Failed to write {SUNO_LOGIN_PS1}: {e}", file=sys.stderr)


def launch_chrome_login_via_powershell(chrome_path: Path) -> None:
    _ensure_suno_login_ps1_exists()
    PERSIST_PROFILE_DIR.mkdir(parents=True, exist_ok=True)

    creationflags = 0
    try:
        creationflags = subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP
    except Exception:
        creationflags = 0

    cmd = [
        "powershell.exe",
        "-NoProfile",
        "-ExecutionPolicy", "Bypass",
        "-File", str(SUNO_LOGIN_PS1),
        "-ChromePath", str(chrome_path),
        "-ProfileDir", str(PERSIST_PROFILE_DIR),
        "-Url", "https://suno.com/",
    ]

    print("[LOGIN] Launching Chrome via PowerShell:")
    print("        " + " ".join(cmd))

    subprocess.Popen(
        cmd,
        stdin=subprocess.DEVNULL,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        creationflags=creationflags
    )


def _profile_looks_initialized(profile_dir: Path) -> bool:
    try:
        if (profile_dir / "Default").exists():
            return True
        if (profile_dir / "Local State").exists():
            return True
    except Exception:
        return False
    return False


def first_launch_bootstrap_then_exit_if_needed():
    chrome_path = ensure_chrome_exe()
    if not chrome_path or not chrome_path.exists():
        _bootstrap_message("Chrome not available", "Chrome is required. Install it (or browse to chrome.exe) and rerun.")
        sys.exit(0)

    if BOOTSTRAP_MARK_FILE.exists():
        print("[BOOTSTRAP] Marker exists. Continuing normal launch.")
        return

    need_bootstrap = not _profile_looks_initialized(PERSIST_PROFILE_DIR)

    if not need_bootstrap:
        try:
            PERSIST_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
            BOOTSTRAP_MARK_FILE.write_text("ok", encoding="utf-8")
        except Exception:
            pass
        return

    if not PLAYWRIGHT_AVAILABLE:
        _bootstrap_message(
            "Playwright missing",
            "Playwright is required to initialize the Chrome profile.\n\nInstall:\n"
            "  py -m pip install --upgrade playwright\n"
            "  py -m playwright install"
        )
        sys.exit(0)

    PERSIST_PROFILE_DIR.mkdir(parents=True, exist_ok=True)

    print("[BOOTSTRAP] Initializing profile with Playwright (headless)…")
    try:
        with sync_playwright() as p:  # type: ignore
            ctx = p.chromium.launch_persistent_context(
                user_data_dir=str(PERSIST_PROFILE_DIR),
                executable_path=str(chrome_path),
                headless=True,
                accept_downloads=True,
                args=[
                    "--no-first-run",
                    "--no-default-browser-check",
                    "--disable-notifications",
                ],
            )
            try:
                ctx.close()
            except Exception:
                pass
    except Exception as e:
        _bootstrap_message("Bootstrap failed", f"Failed to initialize profile:\n\n{e}")
        sys.exit(0)

    try:
        BOOTSTRAP_MARK_FILE.write_text("ok", encoding="utf-8")
    except Exception as e:
        print(f"[BOOTSTRAP] Warning: could not write marker: {e}", file=sys.stderr)

    _bootstrap_message(
        "First run: Suno login required",
        "Chrome will open for Suno login.\n\n"
        "1) Sign in to Suno\n"
        "2) Close Chrome completely\n"
        "3) The downloader will relaunch automatically"
    )

    _spawn_watcher_detached()
    launch_chrome_login_via_powershell(chrome_path)
    _hard_exit_soon(250)


# ===================== Utilities =====================

def format_created(raw) -> str:
    try:
        iso = str(raw).rstrip("Z")
        dt = datetime.fromisoformat(iso)
        return dt.strftime("%m-%d-%Y %I:%M:%S %p")
    except Exception:
        return str(raw)


def extract_id(s: str) -> str:
    patterns = [
        re.compile(r"/playlists?/([A-Za-z0-9-]+)"),
        re.compile(r"/playlist/([A-Za-z0-9-]+)"),
        re.compile(r"/clips?/([A-Za-z0-9-]+)"),
        re.compile(r"/song/([A-Za-z0-9-]+)"),
        re.compile(r"[?&]id=([A-Za-z0-9-]+)"),
        re.compile(r"([A-Za-z0-9-]{22,36})"),
    ]
    for pat in patterns:
        m = pat.search(s)
        if m:
            return m.group(1)
    raise ValueError("Could not extract ID from input.")


def sanitize(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "", name or "")


def safe_sheet_name(name: str) -> str:
    n = (name or "Sheet").strip()
    n = re.sub(r"[:\\/?*\[\]]", "", n).strip() or "Sheet"
    if len(n) > 31:
        n = n[:31].rstrip() or "Sheet"
    return n


# ===================== Suno API =====================

def fetch_playlist(playlist_id: str) -> Tuple[str, List[Dict[str, Any]]]:
    clips: List[Dict[str, Any]] = []
    page = 1
    playlist_name = ""
    rel_index = 1

    while True:
        url = f"https://studio-api.prod.suno.com/api/playlist/{playlist_id}/?page={page}"
        r = requests.get(url, timeout=60)
        r.raise_for_status()
        data = r.json()

        if not playlist_name:
            playlist_name = data.get("name", "Playlist")

        batch = data.get("playlist_clips", [])
        if not batch:
            break

        for item in batch:
            clip = item.get("clip", {}) or {}
            md = clip.get("metadata", {}) or {}

            try:
                duration = int(float(md.get("duration", 0) or 0))
            except Exception:
                duration = 0
            m, s = divmod(duration, 60)

            sliders = md.get("control_sliders", {}) or {}
            weight = sliders.get("style_weight")
            creativity = sliders.get("weirdness_constraint")

            lyrics = (
                clip.get("lyrics")
                or md.get("lyrics")
                or md.get("lyric")
                or md.get("prompt")
                or ""
            )

            clips.append({
                "title":      clip.get("title", "") or "",
                "id":         clip.get("id", "") or "",
                "duration":   f"{m}:{s:02d}",
                "tags":       md.get("tags", "") or "",
                "prompt":     md.get("prompt", "") or "",
                "lyrics":     lyrics,
                "gpt":        md.get("gpt_description_prompt", "") or "",
                "type":       md.get("type", "") or "",
                "model":      clip.get("major_model_version", "") or "",
                "model_name": clip.get("model_name", "") or "",
                "weight":     weight,
                "creativity": creativity,
                "img":        clip.get("image_large_url"),
                "aud":        clip.get("audio_url"),
                "vid":        clip.get("video_url"),
                "created":    clip.get("created_at", "") or "",
                "rel_idx":    rel_index,
                "playlist":   playlist_name
            })
            rel_index += 1

        page += 1

    return playlist_name, clips


def fetch_clip(clip_id: str) -> List[Dict[str, Any]]:
    url = f"https://studio-api.prod.suno.com/api/clip/{clip_id}"
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    data = r.json()
    clip = data.get("clip", data) or {}
    md = clip.get("metadata", {}) or {}

    try:
        duration = int(float(md.get("duration", 0) or 0))
    except Exception:
        duration = 0
    m, s = divmod(duration, 60)

    sliders = md.get("control_sliders", {}) or {}
    weight = sliders.get("style_weight")
    creativity = sliders.get("weirdness_constraint")

    lyrics = (
        clip.get("lyrics")
        or md.get("lyrics")
        or md.get("lyric")
        or md.get("prompt")
        or ""
    )

    return [{
        "title":      clip.get("title", clip_id) or clip_id,
        "id":         clip_id,
        "duration":   f"{m}:{s:02d}",
        "tags":       md.get("tags", "") or "",
        "prompt":     md.get("prompt", "") or "",
        "lyrics":     lyrics,
        "gpt":        md.get("gpt_description_prompt", "") or "",
        "type":       md.get("type", "") or "",
        "model":      clip.get("major_model_version", "") or "",
        "model_name": clip.get("model_name", "") or "",
        "weight":     weight,
        "creativity": creativity,
        "img":        clip.get("image_large_url"),
        "aud":        clip.get("audio_url"),
        "vid":        clip.get("video_url"),
        "created":    clip.get("created_at", "") or "",
        "rel_idx":    1,
        "playlist":   "Unsorted"
    }]


# ===================== File download helpers =====================

def save_mp3(clip: Dict[str, Any], folder: str, filename_base: str) -> Optional[str]:
    os.makedirs(folder, exist_ok=True)
    mp3_path = os.path.join(folder, f"{filename_base}.mp3")
    url = clip.get("aud")
    if not url:
        return None
    try:
        r = requests.get(url, timeout=120)
        r.raise_for_status()
        with open(mp3_path, "wb") as f:
            f.write(r.content)
        return mp3_path
    except Exception as e:
        print(f"[MP3] Download failed: {e}", file=sys.stderr)
        return None


def download_image(url: str, folder: str, filename_base: str) -> Tuple[Optional[str], Optional[str]]:
    if not url:
        return None, None
    os.makedirs(folder, exist_ok=True)
    ext = os.path.splitext(urlparse(url).path)[1] or ".jpg"
    full = os.path.join(folder, f"{filename_base}{ext}")
    try:
        r = requests.get(url, timeout=120)
        r.raise_for_status()
        with open(full, "wb") as f:
            f.write(r.content)

        thumbs = os.path.join(folder, "_thumbs")
        os.makedirs(thumbs, exist_ok=True)
        thumb = os.path.join(thumbs, f"{filename_base}.png")

        img = PILImage.open(full)
        img = ImageOps.fit(img, (40, 40), method=PILImage.Resampling.LANCZOS)
        img.save(thumb)
        return full, thumb
    except Exception as e:
        print(f"[IMG] Download/thumb failed: {e}", file=sys.stderr)
        return None, None


def save_txt(folder: str, filename_base: str, text: str) -> Optional[str]:
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, f"{filename_base}.txt")
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(text or "")
        return path
    except Exception as e:
        print(f"[TXT] Save failed: {e}", file=sys.stderr)
        return None


# ===================== Tagging (MP3 + WAV) =====================

def _populate_id3_common(id3: ID3, clip: Dict[str, Any], img_path: Optional[str]):
    # Clear existing frames
    for frame in list(id3.keys()):
        try:
            del id3[frame]
        except Exception:
            pass

    # Core
    id3.add(TIT2(encoding=3, text=str(clip.get("title", ""))))
    id3.add(TALB(encoding=3, text=str(clip.get("playlist", ""))))
    id3.add(TRCK(encoding=3, text=str(clip.get("rel_idx", ""))))

    # Genre/tags
    if clip.get("tags"):
        id3.add(TCON(encoding=3, text=str(clip.get("tags"))))

    # Lyrics
    lyrics_text = clip.get("lyrics") or clip.get("prompt") or ""
    if lyrics_text:
        id3.add(USLT(encoding=3, lang="eng", desc="", text=str(lyrics_text)))

    # Created date
    created_raw = clip.get("created")
    if created_raw:
        id3.add(TCOP(encoding=3, text=format_created(created_raw)))

    # Comment: gpt + sliders
    original_comment = clip.get("gpt", "") or ""
    parts: List[str] = []
    if clip.get("weight") is not None:
        parts.append(f"Weight: {clip['weight']}")
    if clip.get("creativity") is not None:
        parts.append(f"Creativity: {clip['creativity']}")
    comment_text = " | ".join(filter(None, [original_comment, ", ".join(parts)]))
    if comment_text:
        id3.add(COMM(encoding=3, lang="eng", desc="", text=str(comment_text)))

    # ID in WXXX + TXXX
    clip_id = (str(clip.get("id") or "")).strip()
    if clip_id:
        id3.add(WXXX(encoding=3, desc="ID", url=clip_id))
        id3.add(TXXX(encoding=3, desc="ID", text=[clip_id]))

    # Cover art
    if img_path and os.path.exists(img_path):
        ext = os.path.splitext(img_path)[1].lower()
        mime = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
        with open(img_path, "rb") as imgf:
            id3.add(APIC(encoding=3, mime=mime, type=3, desc="Cover", data=imgf.read()))


def embed_tags_full_rewrite_mp3(mp3_path: str, clip: Dict[str, Any], img_path: Optional[str]):
    try:
        try:
            MP3(mp3_path)
        except Exception:
            pass

        try:
            id3 = ID3(mp3_path)
        except ID3NoHeaderError:
            id3 = ID3()
        except Exception:
            id3 = ID3()

        _populate_id3_common(id3, clip, img_path)
        id3.save(mp3_path, v2_version=4)

    except Exception as e:
        print(f"[TAG] Failed tagging MP3 {mp3_path}: {e}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)


def _riff_info_pack_string(s: str) -> bytes:
    if s is None:
        s = ""
    s = str(s)
    b = s.encode("utf-8", errors="replace") + b"\x00"
    return b


def _make_riff_info_chunk(fields: Dict[str, str]) -> bytes:
    # Windows-friendly RIFF INFO tags:
    # INAM (Title), IPRD (Album/Product), ITRK (Track), IGNR (Genre),
    # ICMT (Comments), ICRD (Creation Date)
    mapping = {
        "title": "INAM",
        "album": "IPRD",
        "track": "ITRK",
        "genre": "IGNR",
        "comment": "ICMT",
        "date": "ICRD",
    }

    subchunks: List[bytes] = []
    for k, cid in mapping.items():
        val = fields.get(k, "")
        if not val:
            continue
        data = _riff_info_pack_string(val)
        size = len(data)
        sub = cid.encode("ascii") + struct.pack("<I", size) + data
        if size % 2 == 1:
            sub += b"\x00"
        subchunks.append(sub)

    list_data = b"INFO" + b"".join(subchunks)
    list_size = len(list_data)
    chunk = b"LIST" + struct.pack("<I", list_size) + list_data
    if list_size % 2 == 1:
        chunk += b"\x00"
    return chunk


def _copy_stream(fsrc, fdst, nbytes: int, buf: int = 1024 * 1024) -> None:
    remaining = nbytes
    while remaining > 0:
        take = buf if remaining > buf else remaining
        b = fsrc.read(take)
        if not b:
            break
        fdst.write(b)
        remaining -= len(b)


def _rewrite_wav_remove_info_and_append(wav_path: str, info_chunk: bytes) -> None:
    tmp_path = wav_path + ".tmp"
    with open(wav_path, "rb") as src, open(tmp_path, "wb") as dst:
        header = src.read(12)
        if len(header) != 12 or header[0:4] != b"RIFF" or header[8:12] != b"WAVE":
            raise RuntimeError("Not a valid RIFF/WAVE file")

        dst.write(header)

        while True:
            hdr = src.read(8)
            if len(hdr) < 8:
                break
            cid = hdr[0:4]
            size = struct.unpack("<I", hdr[4:8])[0]

            if cid == b"LIST" and size >= 4:
                list_type = src.read(4)
                remaining = size - 4
                if list_type == b"INFO":
                    src.seek(remaining, os.SEEK_CUR)
                    if size % 2 == 1:
                        src.seek(1, os.SEEK_CUR)
                    continue

                dst.write(hdr)
                dst.write(list_type)
                _copy_stream(src, dst, remaining)
                if size % 2 == 1:
                    pad = src.read(1)
                    if pad:
                        dst.write(pad)
                continue

            dst.write(hdr)
            _copy_stream(src, dst, size)
            if size % 2 == 1:
                pad = src.read(1)
                if pad:
                    dst.write(pad)

        dst.write(info_chunk)

        file_size = dst.tell()
        riff_size = file_size - 8
        dst.seek(4)
        dst.write(struct.pack("<I", riff_size))

    os.replace(tmp_path, wav_path)


def _write_riff_info_tags_for_windows(wav_path: str, clip: Dict[str, Any]) -> None:
    title = str(clip.get("title", "") or "")
    album = str(clip.get("playlist", "") or "")
    track = str(clip.get("rel_idx", "") or "")
    genre = str(clip.get("tags", "") or "")
    created = clip.get("created")
    date = format_created(created) if created else ""

    clip_id = (str(clip.get("id") or "")).strip()
    base_comment = str(clip.get("gpt", "") or "")
    parts: List[str] = []
    if base_comment:
        parts.append(base_comment)
    if clip_id:
        parts.append(f"ID: {clip_id}")
    if clip.get("weight") is not None:
        parts.append(f"Weight: {clip['weight']}")
    if clip.get("creativity") is not None:
        parts.append(f"Creativity: {clip['creativity']}")
    comment = " | ".join([p for p in parts if p])

    info = {
        "title": title,
        "album": album,
        "track": track,
        "genre": genre,
        "comment": comment,
        "date": date,
    }

    chunk = _make_riff_info_chunk(info)
    _rewrite_wav_remove_info_and_append(wav_path, chunk)


def embed_tags_full_rewrite_wav(wav_path: str, clip: Dict[str, Any], img_path: Optional[str]):
    if not WAVE_AVAILABLE:
        print("[WAVTAG] mutagen.wave not available; cannot tag WAV.", file=sys.stderr)
        return
    try:
        w = WAVE(wav_path)  # type: ignore

        try:
            if w.tags is None:
                w.add_tags()
        except Exception:
            pass

        id3 = w.tags
        if id3 is None:
            try:
                id3 = ID3(wav_path)
            except Exception:
                id3 = ID3()

        _populate_id3_common(id3, clip, img_path)

        try:
            w.save(v2_version=4)
        except TypeError:
            w.save()
        except Exception:
            try:
                id3.save(wav_path, v2_version=4)
            except Exception:
                raise

        try:
            _write_riff_info_tags_for_windows(wav_path, clip)
        except Exception as e:
            print(f"[WAVTAG] RIFF INFO write failed (non-fatal): {e}", file=sys.stderr)

    except Exception as e:
        print(f"[WAVTAG] Failed tagging WAV {wav_path}: {e}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)


# ===================== Retag: fill missing only; preserve timestamps =====================

def _get_id_wxxx_frame(id3: ID3):
    try:
        for fr in id3.getall("WXXX"):
            try:
                if str(fr.desc).strip().lower() == "id":
                    return fr
            except Exception:
                pass
    except Exception:
        pass
    return None


def _get_id_txxx_frame(id3: ID3):
    try:
        for fr in id3.getall("TXXX"):
            try:
                if str(fr.desc).strip().lower() == "id":
                    return fr
            except Exception:
                pass
    except Exception:
        pass
    return None


def retag_mp3_fill_missing_preserve_timestamps(mp3_path: str, clip_id: str) -> bool:
    if not clip_id or not os.path.exists(mp3_path):
        return False

    st = os.stat(mp3_path)
    atime, mtime = st.st_atime, st.st_mtime

    changed = False
    try:
        try:
            id3 = ID3(mp3_path)
        except ID3NoHeaderError:
            id3 = ID3()

        if _get_id_wxxx_frame(id3) is None:
            id3.add(WXXX(encoding=3, desc="ID", url=clip_id))
            changed = True
        if _get_id_txxx_frame(id3) is None:
            id3.add(TXXX(encoding=3, desc="ID", text=[clip_id]))
            changed = True

        if changed:
            id3.save(mp3_path, v2_version=4)
        return changed

    finally:
        try:
            os.utime(mp3_path, (atime, mtime))
        except Exception:
            pass


def _vorbis_has_nonempty(tags, key: str) -> bool:
    try:
        if key not in tags:
            return False
        vals = tags.get(key)
        if not vals:
            return False
        return any(str(v).strip() for v in vals)
    except Exception:
        return False


def retag_flac_fill_missing_preserve_timestamps(flac_path: str, clip_id: str) -> bool:
    if not FLAC_AVAILABLE or not clip_id or not os.path.exists(flac_path):
        return False

    st = os.stat(flac_path)
    atime, mtime = st.st_atime, st.st_mtime

    changed = False
    try:
        audio = FLAC(flac_path)  # type: ignore

        if not _vorbis_has_nonempty(audio.tags or {}, "ID"):
            audio["ID"] = [clip_id]
            changed = True

        if changed:
            audio.save()
        return changed

    finally:
        try:
            os.utime(flac_path, (atime, mtime))
        except Exception:
            pass


# ===================== Duplicate detection by ID =====================

def mp3_extract_clip_id(mp3_path: str) -> Optional[str]:
    try:
        id3 = ID3(mp3_path)
    except Exception:
        return None

    try:
        for fr in id3.getall("WXXX"):
            try:
                if str(fr.desc).strip().lower() == "id":
                    val = str(fr.url).strip()
                    if UUID_RE.fullmatch(val):
                        return val
            except Exception:
                pass
    except Exception:
        pass

    try:
        for fr in id3.getall("TXXX"):
            try:
                if str(fr.desc).strip().lower() == "id":
                    txt = fr.text
                    if isinstance(txt, (list, tuple)):
                        txt = " ".join([str(x) for x in txt])
                    txt = str(txt).strip()
                    m = UUID_RE.search(txt)
                    if m:
                        return m.group(0)
            except Exception:
                pass
    except Exception:
        pass

    # Back-compat: if older files already have the ID elsewhere, keep reading it (does not write it)
    try:
        frames = id3.getall("TCOM")
        if frames:
            txt = frames[0].text
            if isinstance(txt, (list, tuple)):
                txt = " ".join([str(x) for x in txt])
            txt = str(txt).strip()
            m = UUID_RE.search(txt)
            if m:
                return m.group(0)
    except Exception:
        pass

    return None


def flac_extract_clip_id(flac_path: str) -> Optional[str]:
    if not FLAC_AVAILABLE:
        return None
    try:
        a = FLAC(flac_path)  # type: ignore
    except Exception:
        return None
    try:
        for key in ("ID", "COMPOSER"):
            vals = a.get(key)
            if vals:
                for v in vals:
                    s = str(v).strip()
                    m = UUID_RE.search(s)
                    if m:
                        return m.group(0)
    except Exception:
        pass
    return None


def scan_audio_dir_ids(audio_dir: str) -> Tuple[Dict[str, Dict[str, Optional[str]]], Dict[str, int]]:
    """
    Returns:
      existing_by_id: dict clip_id -> {"base": stem, "mp3": path|None, "flac": path|None}
      version_seed: dict raw_title -> max_version_seen (from existing "RAW Vn")
    """
    existing_by_id: Dict[str, Dict[str, Optional[str]]] = {}
    stems: List[str] = []

    p = Path(audio_dir)
    if not p.exists():
        return existing_by_id, {}

    try:
        for f in p.glob("*.mp3"):
            cid = mp3_extract_clip_id(str(f))
            if cid:
                existing_by_id.setdefault(cid, {"base": f.stem, "mp3": str(f), "flac": None})
                stems.append(f.stem)

        for f in p.glob("*.flac"):
            cid = flac_extract_clip_id(str(f))
            if cid:
                entry = existing_by_id.setdefault(cid, {"base": f.stem, "mp3": None, "flac": str(f)})
                entry["flac"] = str(f)
                stems.append(f.stem)

    except Exception as e:
        print(f"[DUP] scan_audio_dir_ids failed: {e}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)

    version_seed: Dict[str, int] = {}
    mver = re.compile(r"^(.*)\s+V(\d+)$", re.IGNORECASE)
    for stem in stems:
        m = mver.match(stem)
        if not m:
            continue
        raw = m.group(1)
        try:
            v = int(m.group(2))
        except Exception:
            continue
        prev = version_seed.get(raw, 0)
        if v > prev:
            version_seed[raw] = v

    return existing_by_id, version_seed


def unique_clips_by_id(clips: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    seen: Set[str] = set()
    for c in clips:
        cid = str(c.get("id") or "").strip()
        if cid and cid in seen:
            continue
        if cid:
            seen.add(cid)
        out.append(c)
    return out


# ===================== XLSX index helpers (preserve thumbnails) =====================

def header_col_map(ws) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            out[str(cell.value)] = idx
    return out


def _open_wb_and_sheet(xlsx_path: str, sheet_name: str):
    wb = load_workbook(xlsx_path)
    safe = safe_sheet_name(sheet_name)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif safe in wb.sheetnames:
        ws = wb[safe]
    else:
        ws = wb.active
    return wb, ws


def update_index_titles_in_place(xlsx_path: str, sheet_name: str, id_to_title: Dict[str, str]) -> bool:
    if not os.path.exists(xlsx_path):
        return False
    wb, ws = _open_wb_and_sheet(xlsx_path, sheet_name)

    cols = header_col_map(ws)
    if "ID" not in cols or "Title" not in cols:
        return False

    id_col = cols["ID"]
    title_col = cols["Title"]

    changed = False
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=id_col).value
        if v is None:
            continue
        sid = str(v)
        if sid in id_to_title:
            new_title = id_to_title[sid]
            cell = ws.cell(row=r, column=title_col)
            if cell.value != new_title:
                cell.value = new_title
                changed = True

    if changed:
        wb.save(xlsx_path)
    return changed


def append_missing_rows_preserve_thumbs(xlsx_path: str, sheet_name: str, clips: List[Dict[str, Any]]) -> bool:
    if not os.path.exists(xlsx_path):
        return False
    wb, ws = _open_wb_and_sheet(xlsx_path, sheet_name)

    cols = header_col_map(ws)
    if "ID" not in cols:
        return False
    id_col = cols["ID"]

    existing_ids = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=id_col).value
        if v is not None:
            existing_ids.add(str(v))

    to_add = [c for c in clips if c.get("id") and str(c["id"]) not in existing_ids]
    if not to_add:
        return False

    for c in to_add:
        ws.append([
            c.get("master_idx"),
            c.get("rel_idx"),
            c.get("playlist"),
            c.get("index_title") or c.get("title"),
            c.get("duration"),
            None,
            c.get("tags"),
            c.get("gpt"),
            c.get("lyrics") or c.get("prompt"),
            c.get("id"),
            format_created(c.get("created")),
            c.get("model"),
            c.get("model_name"),
            c.get("type"),
            c.get("weight") if c.get("weight") is not None else "N/A",
            c.get("creativity") if c.get("creativity") is not None else "N/A",
            c.get("img"),
            c.get("aud"),
            c.get("vid"),
        ])
        row = ws.max_row
        try:
            ws.row_dimensions[row].height = 30
        except Exception:
            pass

        thumb = c.get("thumb_path")
        if thumb and os.path.exists(thumb):
            try:
                ws.add_image(XLImage(thumb), f"F{row}")
            except Exception as e:
                print(f"[XLSX] Failed to add thumbnail for row {row}: {e}", file=sys.stderr)

    wb.save(xlsx_path)
    return True


def create_index_xlsx(clips: List[Dict[str, Any]], xlsx_path: str, sheet_title: str):
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name(sheet_title)
    ws.append(HEADERS)

    for clip in clips:
        clip["_row"] = ws.max_row + 1
        ws.append([
            clip.get("master_idx"),
            clip.get("rel_idx"),
            clip.get("playlist"),
            clip.get("index_title") or clip.get("title"),
            clip.get("duration"),
            None,
            clip.get("tags"),
            clip.get("gpt"),
            clip.get("lyrics") or clip.get("prompt"),
            clip.get("id"),
            format_created(clip.get("created")),
            clip.get("model"),
            clip.get("model_name"),
            clip.get("type"),
            clip.get("weight") if clip.get("weight") is not None else "N/A",
            clip.get("creativity") if clip.get("creativity") is not None else "N/A",
            clip.get("img"),
            clip.get("aud"),
            clip.get("vid"),
        ])

    for r in range(1, ws.max_row + 1):
        try:
            ws.row_dimensions[r].height = 30
        except Exception:
            pass

    for clip in clips:
        thumb = clip.get("thumb_path")
        if thumb and os.path.exists(thumb):
            try:
                ws.add_image(XLImage(thumb), f"F{clip['_row']}")
            except Exception as e:
                print(f"[XLSX] Failed to add thumbnail: {e}", file=sys.stderr)

    wb.save(xlsx_path)


# ===================== Playwright WAV UI automation =====================

def pw_wait_until(fn, timeout_s: float = 15.0, poll_s: float = 0.25) -> bool:
    start = time.time()
    while time.time() - start < timeout_s:
        try:
            if fn():
                return True
        except Exception:
            pass
        time.sleep(poll_s)
    return False


def pw_download_entry_visible(page: "Page") -> bool:
    try:
        loc = page.locator('button.context-menu-button:has-text("Download")').first
        return loc.count() > 0 and loc.is_visible()
    except Exception:
        return False


def pw_find_more_menu_candidates(page: "Page"):
    return [
        page.locator('button[aria-label="More Menu Options"]'),
        page.locator('button[title="More Menu Options"]'),
        page.locator('button[aria-label*="More" i]'),
        page.locator(f'button:has(svg path[d^="{MORE_MENU_PATH_PREFIX}"])'),
        page.locator(f'[role="button"]:has(svg path[d^="{MORE_MENU_PATH_PREFIX}"])'),
    ]


def pw_ensure_menu_open(page: "Page") -> None:
    if pw_download_entry_visible(page):
        return

    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(150)
    except Exception:
        pass

    page.wait_for_timeout(800)

    for cand in pw_find_more_menu_candidates(page):
        try:
            n = cand.count()
        except Exception:
            n = 0
        if n <= 0:
            continue

        for i in range(min(n, 12)):
            el = cand.nth(i)
            try:
                el.scroll_into_view_if_needed()
                if not el.is_visible():
                    continue
                el.click()
                page.wait_for_timeout(200)
                if pw_download_entry_visible(page):
                    return
            except Exception:
                pass

    try:
        main = page.locator("main").first
        if main.count():
            main.click(button="right")
            page.wait_for_timeout(250)
            if pw_download_entry_visible(page):
                return
    except Exception:
        pass

    raise RuntimeError("Could not open hidden menu (Download entry never became visible).")


def pw_wait_for_modal_download_button_enabled(page: "Page"):
    page.locator('text=/Download WAV Audio/i').first.wait_for(timeout=15_000)

    dl_btn = page.get_by_role("button", name=re.compile(r"Download File", re.I))
    dl_btn.wait_for(state="visible", timeout=15_000)

    ok = pw_wait_until(lambda: dl_btn.is_enabled(), timeout_s=60.0, poll_s=0.5)
    if not ok:
        raise RuntimeError("Download File never became enabled (still preparing or blocked).")
    return dl_btn


def pw_trigger_ui_wav_download(page: "Page", per_song_timeout_ms: int = PW_PER_SONG_TIMEOUT_MS):
    pw_ensure_menu_open(page)

    download_btn = page.locator('button.context-menu-button:has-text("Download")').first
    download_btn.hover()
    page.wait_for_timeout(150)

    wav_entry = page.locator(
        '[role="menu"] :is(button,[role="menuitem"]):has-text("WAV Audio"), '
        ':is(button,[role="menuitem"]):has-text("WAV Audio")'
    ).first

    try:
        wav_entry.wait_for(state="visible", timeout=5000)
        wav_entry.click()
    except PWTimeoutError:
        download_btn.hover()
        page.wait_for_timeout(150)
        wav_entry.wait_for(state="visible", timeout=5000)
        wav_entry.click()

    dl_file_btn = pw_wait_for_modal_download_button_enabled(page)

    with page.expect_download(timeout=per_song_timeout_ms) as dlinfo:
        dl_file_btn.click()

    return dlinfo.value


def pw_login_gate_detected(page: "Page") -> bool:
    try:
        page.wait_for_timeout(600)
        sign_in = page.locator(r'text=/\b(sign in|log in)\b/i').first
        return bool(sign_in.count() and sign_in.is_visible())
    except Exception:
        return False


# ===================== GUI App =====================

class SunoDownloader(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Suno Downloader")
        self.geometry("1230x860")
        self.configure(bg=BG_COLOR, highlightthickness=0)

        self.custom_font = font.Font(family=FONT_FAMILY, size=FONT_SIZE)
        self.icon_font = font.Font(family="Segoe UI Symbol", size=13)

        self.pause_event = threading.Event()
        self.stop_event = threading.Event()
        self._worker_thread = None

        self._log_line_count = 0

        tk.Label(
            self,
            text="Enter playlist/clip URLs or IDs (one per line):",
            bg=BG_COLOR,
            fg=FG_COLOR,
            font=self.custom_font
        ).pack(anchor="w", padx=10, pady=(10, 0))

        self.text_box = tk.Text(
            self,
            width=160,
            height=6,
            bg=ENTRY_BG,
            fg=ENTRY_FG,
            font=self.custom_font
        )
        self.text_box.pack(padx=10, pady=(0, 10))

        opts = tk.Frame(self, bg=BG_COLOR)
        opts.pack(fill=tk.X, padx=10, pady=(0, 10))

        self.audio_var = tk.BooleanVar(value=True)
        self.artwork_var = tk.BooleanVar(value=True)
        self.lyrics_var = tk.BooleanVar(value=True)
        self.genres_var = tk.BooleanVar(value=True)
        self.prompts_var = tk.BooleanVar(value=True)
        self.pl_idx_var = tk.BooleanVar(value=True)
        self.master_idx_var = tk.BooleanVar(value=True)
        self.wav_var = tk.BooleanVar(value=False)
        self.retag_only_var = tk.BooleanVar(value=False)

        labels = [
            ("Audio (MP3)", self.audio_var),
            ("Artwork", self.artwork_var),
            ("Lyrics", self.lyrics_var),
            ("Genres", self.genres_var),
            ("Prompts", self.prompts_var),
            ("Playlist Index", self.pl_idx_var),
            ("Master Index", self.master_idx_var),
            ("WAV", self.wav_var),
            ("Retag Only", self.retag_only_var),
        ]

        self.checkbuttons = []
        for i, (txt, var) in enumerate(labels):
            r = 0 if i < 5 else 1
            c = i if i < 5 else i - 5
            cb = tk.Checkbutton(
                opts,
                text=txt,
                variable=var,
                bg=BG_COLOR,
                fg=CHECK_TEXT_COLOR,
                selectcolor=SELECT_COLOR,
                activeforeground=CHECK_TEXT_COLOR,
                activebackground=BG_COLOR,
                disabledforeground=DISABLED_COLOR,
                font=self.custom_font,
                command=self.update_option_states
            )
            cb.grid(row=r, column=c, sticky="w", padx=6, pady=2)
            self.checkbuttons.append(cb)

        self._cb_audio = 0
        self._cb_art = 1
        self._cb_lyrics = 2
        self._cb_genres = 3
        self._cb_prompts = 4
        self._cb_plidx = 5
        self._cb_master = 6
        self._cb_wav = 7
        self._cb_retag = 8

        if not PLAYWRIGHT_AVAILABLE:
            self.checkbuttons[self._cb_wav].config(state=tk.DISABLED)
            self.wav_var.set(False)

        btns = tk.Frame(self, bg=BG_COLOR)
        btns.pack(fill=tk.X, padx=10)

        self.load_btn = tk.Button(
            btns,
            text="Load URLs from File",
            command=self.load_from_file,
            bg=BUTTON_BG,
            fg=BUTTON_FG,
            font=self.custom_font
        )
        self.load_btn.pack(side=tk.LEFT)

        self.load_session_btn = tk.Button(
            btns,
            text="Load Session",
            command=self.load_session,
            bg=BUTTON_BG,
            fg=BUTTON_FG,
            font=self.custom_font
        )
        self.load_session_btn.pack(side=tk.LEFT, padx=(8, 0))

        self.save_session_btn = tk.Button(
            btns,
            text="Save Session",
            command=self.save_session,
            bg=BUTTON_BG,
            fg=BUTTON_FG,
            font=self.custom_font
        )
        self.save_session_btn.pack(side=tk.LEFT, padx=(8, 0))

        self.login_btn = tk.Button(
            btns,
            text="Suno Login (Chrome)",
            command=self.login_and_relaunch,
            bg=BUTTON_BG,
            fg=BUTTON_FG,
            font=self.custom_font
        )
        self.login_btn.pack(side=tk.RIGHT)

        self.download_btn = tk.Button(
            btns,
            text="Run",
            command=self.start_download,
            bg=BUTTON_BG,
            fg=BUTTON_FG,
            font=self.custom_font
        )
        self.download_btn.pack(side=tk.RIGHT, padx=(0, 8))

        self.progress = Progressbar(self, orient=tk.HORIZONTAL, length=1190, mode="determinate")
        self.progress.pack(pady=5)

        self.progress_label = tk.Label(
            self,
            text="0/0",
            bg=BG_COLOR,
            fg=CHECK_TEXT_COLOR,
            font=self.custom_font
        )
        self.progress_label.pack()

        controls = tk.Frame(self, bg=BG_COLOR)
        controls.pack(fill=tk.X, padx=10, pady=(4, 4))

        self.btn_clear = tk.Button(
            controls, text="🧹", font=self.icon_font, width=3,
            command=self.clear_console, bg=BUTTON_BG, fg=BUTTON_FG
        )
        self.btn_clear.pack(side=tk.LEFT, padx=(0, 6))

        self.btn_pause = tk.Button(
            controls, text="⏸", font=self.icon_font, width=3,
            command=self.toggle_pause, bg=BUTTON_BG, fg=BUTTON_FG, state=tk.DISABLED
        )
        self.btn_pause.pack(side=tk.LEFT, padx=(0, 6))

        self.btn_stop = tk.Button(
            controls, text="⏹", font=self.icon_font, width=3,
            command=self.request_stop, bg=BUTTON_BG, fg=BUTTON_FG, state=tk.DISABLED
        )
        self.btn_stop.pack(side=tk.LEFT, padx=(0, 6))

        console_frame = tk.Frame(self, bg=BG_COLOR)
        console_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 8))

        self.console_text = tk.Text(
            console_frame,
            wrap="none",
            bg="#0e0010",
            fg="#c7ffca",
            insertbackground="#c7ffca",
            font=("Consolas", 10),
            height=16
        )
        self.console_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        yscroll = tk.Scrollbar(console_frame, orient="vertical", command=self.console_text.yview)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.console_text.configure(yscrollcommand=yscroll.set)

        self.status_label = tk.Label(
            self,
            text="",
            bg=BG_COLOR,
            fg=CHECK_TEXT_COLOR,
            font=self.custom_font,
            anchor="w"
        )
        self.status_label.pack(fill=tk.X, padx=20, pady=(0, 10))

        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.update_option_states()
        self.after(50, self._drain_log_queue)

    # ---------- session save/load ----------

    def _session_payload(self) -> Dict[str, Any]:
        return {
            "version": 1,
            "saved_at": datetime.now().isoformat(timespec="seconds"),
            "inputs": self.text_box.get("1.0", tk.END).rstrip("\n"),
            "options": {
                "audio": bool(self.audio_var.get()),
                "artwork": bool(self.artwork_var.get()),
                "lyrics": bool(self.lyrics_var.get()),
                "genres": bool(self.genres_var.get()),
                "prompts": bool(self.prompts_var.get()),
                "playlist_index": bool(self.pl_idx_var.get()),
                "master_index": bool(self.master_idx_var.get()),
                "wav": bool(self.wav_var.get()),
                "retag_only": bool(self.retag_only_var.get()),
            },
        }

    def save_session(self):
        try:
            SESSION_DIR.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

        name = simpledialog.askstring("Save Session", "Session name:", parent=self)
        if not name:
            return

        safe = sanitize(name).strip() or "session"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = SESSION_DIR / f"{safe}_{ts}.json"

        payload = self._session_payload()
        try:
            path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
            self._set_status(f"Session saved: {path.name}")
            print(f"[SESSION] Saved: {path}")
        except Exception as e:
            print(f"[SESSION] Save failed: {e}", file=sys.stderr)
            messagebox.showerror("Save failed", str(e))

    def load_session(self):
        try:
            SESSION_DIR.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

        path = filedialog.askopenfilename(
            title="Load Session",
            initialdir=str(SESSION_DIR),
            filetypes=[("Session JSON", "*.json"), ("All files", "*.*")]
        )
        if not path:
            return

        try:
            data = json.loads(Path(path).read_text(encoding="utf-8", errors="replace"))
            inputs = str(data.get("inputs", "") or "")
            opts = data.get("options", {}) or {}

            self.text_box.delete("1.0", tk.END)
            self.text_box.insert(tk.END, inputs)

            self.audio_var.set(bool(opts.get("audio", True)))
            self.artwork_var.set(bool(opts.get("artwork", True)))
            self.lyrics_var.set(bool(opts.get("lyrics", True)))
            self.genres_var.set(bool(opts.get("genres", True)))
            self.prompts_var.set(bool(opts.get("prompts", True)))
            self.pl_idx_var.set(bool(opts.get("playlist_index", True)))
            self.master_idx_var.set(bool(opts.get("master_index", True)))
            self.wav_var.set(bool(opts.get("wav", False)))
            self.retag_only_var.set(bool(opts.get("retag_only", False)))

            self.update_option_states()
            self._set_status(f"Session loaded: {Path(path).name}")
            print(f"[SESSION] Loaded: {path}")
        except Exception as e:
            print(f"[SESSION] Load failed: {e}", file=sys.stderr)
            print(traceback.format_exc(), file=sys.stderr)
            messagebox.showerror("Load failed", str(e))

    # ---------- login button ----------

    def login_and_relaunch(self):
        if self._worker_thread and self._worker_thread.is_alive():
            messagebox.showerror("Busy", "Stop the current job before launching login.")
            return

        chrome_path = ensure_chrome_exe()
        if not chrome_path or not chrome_path.exists():
            return

        _spawn_watcher_detached()
        launch_chrome_login_via_powershell(chrome_path)

        try:
            self.stop_event.set()
            self.pause_event.clear()
        except Exception:
            pass

        _hard_exit_soon(250)

    # ---------- UI helpers ----------

    def _set_status(self, text: str):
        print(f"[STATUS] {text}")
        self.after(0, lambda t=text: self.status_label.config(text=t))

    def _set_running_controls(self, running: bool):
        self.btn_pause.config(state=(tk.NORMAL if running else tk.DISABLED))
        self.btn_stop.config(state=(tk.NORMAL if running else tk.DISABLED))
        if not running:
            self.pause_event.clear()
            self.btn_pause.config(text="⏸")

    def update_option_states(self):
        retag = self.retag_only_var.get()

        audio_cb = self.checkbuttons[self._cb_audio]
        art_cb = self.checkbuttons[self._cb_art]
        lyr_cb = self.checkbuttons[self._cb_lyrics]
        gen_cb = self.checkbuttons[self._cb_genres]
        prm_cb = self.checkbuttons[self._cb_prompts]
        wav_cb = self.checkbuttons[self._cb_wav]

        if retag:
            for cb in (audio_cb, art_cb, lyr_cb, gen_cb, prm_cb, wav_cb):
                cb.config(state=tk.DISABLED)
            self.wav_var.set(False)
        else:
            audio_cb.config(state=tk.NORMAL)
            for cb in (art_cb, lyr_cb, gen_cb, prm_cb):
                cb.config(state=tk.NORMAL)

            if PLAYWRIGHT_AVAILABLE:
                wav_cb.config(state=tk.NORMAL)
            else:
                wav_cb.config(state=tk.DISABLED)
                self.wav_var.set(False)

        vars_list = [
            self.audio_var, self.artwork_var,
            self.lyrics_var, self.genres_var, self.prompts_var,
            self.pl_idx_var, self.master_idx_var, self.wav_var,
            self.retag_only_var
        ]
        for cb, var in zip(self.checkbuttons, vars_list):
            cb.config(
                fg=CHECK_TEXT_COLOR if var.get() else DISABLED_COLOR,
                selectcolor=SELECT_COLOR if var.get() else BG_COLOR
            )

    def load_from_file(self):
        path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt"), ("All files", "*.*")])
        if path:
            with open(path, encoding="utf-8", errors="ignore") as f:
                data = f.read()
            self.text_box.delete("1.0", tk.END)
            self.text_box.insert(tk.END, data)

    # ---------- console ----------

    def clear_console(self):
        try:
            while True:
                UI_LOG_QUEUE.get_nowait()
        except Exception:
            pass
        self.console_text.delete("1.0", tk.END)
        self._log_line_count = 0

    def _append_console(self, text: str):
        try:
            at_bottom = (self.console_text.yview()[1] >= 0.999)
        except Exception:
            at_bottom = True

        self.console_text.insert(tk.END, text)
        self._log_line_count += text.count("\n")

        if self._log_line_count > UI_LOG_MAX_LINES:
            excess = self._log_line_count - UI_LOG_MAX_LINES
            try:
                self.console_text.delete("1.0", f"{excess + 1}.0")
                self._log_line_count -= excess
            except Exception:
                self.console_text.delete("1.0", tk.END)
                self._log_line_count = 0

        if at_bottom:
            try:
                self.console_text.see(tk.END)
            except Exception:
                pass

    def _drain_log_queue(self):
        try:
            for _ in range(500):
                try:
                    line = UI_LOG_QUEUE.get_nowait()
                except Exception:
                    break
                self._append_console(line)
        except Exception:
            pass
        finally:
            self.after(50, self._drain_log_queue)

    # ---------- pause / stop ----------

    def toggle_pause(self):
        if not (self._worker_thread and self._worker_thread.is_alive()):
            return
        if self.pause_event.is_set():
            self.pause_event.clear()
            self.btn_pause.config(text="⏸")
            self._set_status("Resumed.")
        else:
            self.pause_event.set()
            self.btn_pause.config(text="▶")
            self._set_status("Paused.")

    def request_stop(self):
        if not (self._worker_thread and self._worker_thread.is_alive()):
            return
        self.stop_event.set()
        self.pause_event.clear()
        self.btn_pause.config(text="⏸")
        self._set_status("Stop requested…")

    def _wait_if_paused_or_stopped(self) -> bool:
        while self.pause_event.is_set():
            if self.stop_event.is_set():
                return False
            time.sleep(0.1)
        return not self.stop_event.is_set()

    # ---------- WAV downloads ----------

    def download_wavs_in_playwright(self, wav_jobs: List[Dict[str, Any]], base_folder: str):
        if not PLAYWRIGHT_AVAILABLE:
            raise RuntimeError("playwright is not installed")

        chrome_path = ensure_chrome_exe()
        if not chrome_path or not chrome_path.exists():
            raise RuntimeError("Chrome not available (chrome.exe not found)")

        PERSIST_PROFILE_DIR.mkdir(parents=True, exist_ok=True)

        staging_dir = Path(base_folder) / "_pw_downloads_tmp"
        staging_dir.mkdir(parents=True, exist_ok=True)

        self._set_status("WAV: launching Chrome (Playwright)…")

        with sync_playwright() as p:  # type: ignore
            context: "BrowserContext" = p.chromium.launch_persistent_context(
                user_data_dir=str(PERSIST_PROFILE_DIR),
                executable_path=str(chrome_path),
                headless=False,
                accept_downloads=True,
                downloads_path=str(staging_dir),
                args=PW_CHROME_ARGS,
            )
            context.set_default_timeout(60_000)

            page = context.pages[0] if context.pages else context.new_page()

            self._set_status("WAV: opening suno.com…")
            page.goto("https://suno.com/", wait_until="domcontentloaded")
            page.wait_for_timeout(1200)

            if pw_login_gate_detected(page):
                try:
                    context.close()
                except Exception:
                    pass
                raise RuntimeError(
                    "Suno appears signed out in the Chrome profile.\n\n"
                    "Use “Suno Login (Chrome)”, sign in, close Chrome, then try again."
                )

            total = len(wav_jobs)
            failures = 0

            for idx, job in enumerate(wav_jobs, start=1):
                if not self._wait_if_paused_or_stopped():
                    print("[WAV] Stopped by user.")
                    break

                song_url = job["song_url"]
                out_wav = Path(job["out_wav"])
                out_wav.parent.mkdir(parents=True, exist_ok=True)

                if out_wav.exists():
                    print(f"[WAV] {idx}/{total} SKIP (exists): {out_wav.name}")
                    continue

                self._set_status(f"WAV: {idx}/{total} opening song page…")
                print(f"[WAV] {idx}/{total} OPEN: {song_url}")

                try:
                    page.goto(song_url, wait_until="domcontentloaded")
                    page.wait_for_timeout(1600)

                    if not self._wait_if_paused_or_stopped():
                        print("[WAV] Stopped by user.")
                        break

                    self._set_status(f"WAV: {idx}/{total} triggering UI download…")
                    download = pw_trigger_ui_wav_download(page, per_song_timeout_ms=PW_PER_SONG_TIMEOUT_MS)

                    download.save_as(str(out_wav))
                    print(f"[WAV] {idx}/{total} SAVED: {out_wav.name}")

                    clip = job.get("clip") or {}
                    cover_path = job.get("cover_path")
                    if clip:
                        embed_tags_full_rewrite_wav(str(out_wav), clip, cover_path)

                    try:
                        download.delete()
                    except Exception:
                        try:
                            pth = download.path()
                            if pth:
                                Path(pth).unlink(missing_ok=True)
                        except Exception:
                            pass

                    try:
                        page.keyboard.press("Escape")
                    except Exception:
                        pass

                except Exception as e:
                    failures += 1
                    print(f"[WAV] {idx}/{total} FAILED: {e}", file=sys.stderr)
                    print(traceback.format_exc(), file=sys.stderr)
                    try:
                        page.goto("https://suno.com/", wait_until="domcontentloaded")
                        page.wait_for_timeout(800)
                    except Exception:
                        pass

                time.sleep(0.25)

            try:
                context.close()
            except Exception:
                pass

            if failures and not self.stop_event.is_set():
                raise RuntimeError(f"WAV downloads finished with {failures} failure(s).")

    # ---------- main run ----------

    def start_download(self):
        lines = [l.strip() for l in self.text_box.get("1.0", tk.END).splitlines() if l.strip()]
        if not lines:
            messagebox.showerror("Error", "Please enter at least one URL or ID.")
            return

        folder = filedialog.askdirectory(title="Select Base Folder")
        if not folder:
            return

        opts = (
            self.audio_var.get(),
            self.artwork_var.get(),
            self.lyrics_var.get(),
            self.genres_var.get(),
            self.prompts_var.get(),
            self.pl_idx_var.get(),
            self.master_idx_var.get(),
            self.wav_var.get(),
            self.retag_only_var.get(),
        )

        self.stop_event.clear()
        self.pause_event.clear()
        self._set_running_controls(True)

        self.load_btn.config(state=tk.DISABLED)
        self.load_session_btn.config(state=tk.DISABLED)
        self.save_session_btn.config(state=tk.DISABLED)
        self.login_btn.config(state=tk.DISABLED)
        self.download_btn.config(state=tk.DISABLED)

        self._worker_thread = threading.Thread(
            target=self.download_worker,
            args=(lines, folder) + opts,
            daemon=True
        )
        self._worker_thread.start()

    def download_worker(self, lines, folder,
                        do_audio, do_art, do_lyrics, do_genres,
                        do_prompts, do_pl_idx, do_master, do_wav,
                        retag_only):
        try:
            all_clips: List[Dict[str, Any]] = []
            wav_jobs: List[Dict[str, Any]] = []
            total = 0

            for line in lines:
                if not self._wait_if_paused_or_stopped():
                    break
                try:
                    pid = extract_id(line)
                    _, clips = fetch_playlist(pid)
                except Exception:
                    clips = fetch_clip(extract_id(line))
                total += len(clips)

            self.after(0, lambda: self.progress.config(maximum=max(total, 1), value=0))
            self.after(0, lambda: self.progress_label.config(text=f"0/{total}"))

            done = 0
            master_counter = 0
            processed_playlists: Set[str] = set()

            effective_do_audio = bool(do_audio) and not retag_only
            effective_do_wav = bool(do_wav) and not retag_only

            for line in lines:
                if not self._wait_if_paused_or_stopped():
                    break

                try:
                    pid = extract_id(line)
                    self._set_status(f"Fetching metadata for {pid}…")

                    try:
                        name, clips = fetch_playlist(pid)
                    except requests.HTTPError:
                        clips = fetch_clip(pid)
                        name = "Unsorted"

                    clips = unique_clips_by_id(clips)

                    dest = os.path.join(folder, sanitize(name))
                    audio_dir = os.path.join(dest, "Audio")
                    art_dir = os.path.join(dest, "Art")

                    if not retag_only:
                        for sec in ["Audio", "Art", "Lyrics", "Prompt", "Genres"]:
                            os.makedirs(os.path.join(dest, sec), exist_ok=True)
                    else:
                        os.makedirs(audio_dir, exist_ok=True)

                    existing_by_id, version_seed = scan_audio_dir_ids(audio_dir)
                    versions = dict(version_seed)

                    for clip in clips:
                        master_counter += 1
                        clip["master_idx"] = master_counter

                        clip_id = str(clip.get("id") or "").strip()

                        if clip_id and clip_id in existing_by_id and existing_by_id[clip_id].get("base"):
                            file_base = str(existing_by_id[clip_id]["base"])
                        else:
                            raw = sanitize(str(clip.get("title", ""))).strip()
                            if not raw:
                                raw = sanitize(clip_id or "Untitled").strip() or "Untitled"
                            cnt = versions.get(raw, 0) + 1
                            versions[raw] = cnt
                            file_base = f"{raw} V{cnt}"

                        clip["index_title"] = file_base
                        clip["title"] = file_base

                        thumb_guess = os.path.join(art_dir, "_thumbs", f"{file_base}.png")
                        if os.path.exists(thumb_guess):
                            clip["thumb_path"] = thumb_guess

                        if not clip.get("cover_path"):
                            for ext in (".jpg", ".jpeg", ".png", ".webp"):
                                cover_guess = os.path.join(art_dir, f"{file_base}{ext}")
                                if os.path.exists(cover_guess):
                                    clip["cover_path"] = cover_guess
                                    break

                    for clip in clips:
                        if not self._wait_if_paused_or_stopped():
                            break

                        file_base = str(clip["index_title"])
                        clip_id = str(clip.get("id") or "").strip()

                        mp3_expected = os.path.join(audio_dir, f"{file_base}.mp3")
                        flac_expected = os.path.join(audio_dir, f"{file_base}.flac")
                        wav_expected = os.path.join(audio_dir, f"{file_base}.wav")

                        mp3_existing = existing_by_id.get(clip_id, {}).get("mp3") if clip_id else None
                        flac_existing = existing_by_id.get(clip_id, {}).get("flac") if clip_id else None

                        if retag_only:
                            if clip_id:
                                mp3_target = mp3_existing or (mp3_expected if os.path.exists(mp3_expected) else None)
                                if mp3_target and os.path.exists(mp3_target):
                                    self._set_status(f"Retag MP3 (fill missing): {os.path.basename(mp3_target)}")
                                    changed = retag_mp3_fill_missing_preserve_timestamps(mp3_target, clip_id)
                                    print(f"[RETAG] MP3 {os.path.basename(mp3_target)} changed={changed}")

                                flac_target = flac_existing or (flac_expected if os.path.exists(flac_expected) else None)
                                if flac_target and os.path.exists(flac_target):
                                    if FLAC_AVAILABLE:
                                        self._set_status(f"Retag FLAC (fill missing): {os.path.basename(flac_target)}")
                                        changed = retag_flac_fill_missing_preserve_timestamps(flac_target, clip_id)
                                        print(f"[RETAG] FLAC {os.path.basename(flac_target)} changed={changed}")
                                    else:
                                        print("[RETAG] FLAC support not available (mutagen.flac import failed).")
                        else:
                            need_thumb = (do_art or do_pl_idx or do_master or effective_do_audio or effective_do_wav) and not clip.get("thumb_path")
                            need_cover = (effective_do_audio or effective_do_wav) and not clip.get("cover_path")

                            if (need_thumb or need_cover) and clip.get("img"):
                                self._set_status(f"Downloading artwork: {file_base}")
                                full, thumb = download_image(str(clip.get("img")), art_dir, file_base)
                                if full:
                                    clip["cover_path"] = full
                                if thumb:
                                    clip["thumb_path"] = thumb

                            already_have_audio_for_id = bool(
                                clip_id and clip_id in existing_by_id and (existing_by_id[clip_id].get("mp3") or existing_by_id[clip_id].get("flac"))
                            )

                            if effective_do_audio:
                                if already_have_audio_for_id and mp3_existing and os.path.exists(mp3_existing):
                                    print(f"[DUP] Audio exists for ID {clip_id}, skip MP3 download: {Path(mp3_existing).name}")
                                elif os.path.exists(mp3_expected):
                                    print(f"[DUP] MP3 exists, skip: {Path(mp3_expected).name}")
                                else:
                                    self._set_status(f"Downloading MP3: {file_base}.mp3")
                                    mp3_path = save_mp3(clip, audio_dir, file_base)
                                    if mp3_path:
                                        self._set_status(f"Tagging MP3: {file_base}.mp3")
                                        embed_tags_full_rewrite_mp3(mp3_path, clip, clip.get("cover_path"))
                                    else:
                                        print(f"[MP3] Could not download MP3 for {file_base}")

                            if do_prompts:
                                save_txt(os.path.join(dest, "Prompt"), file_base, str(clip.get("gpt", "") or ""))
                            if do_lyrics:
                                save_txt(os.path.join(dest, "Lyrics"), file_base, str(clip.get("lyrics") or clip.get("prompt") or ""))
                            if do_genres:
                                save_txt(os.path.join(dest, "Genres"), file_base, str(clip.get("tags", "") or ""))

                            if effective_do_wav and clip_id:
                                if not os.path.exists(wav_expected):
                                    wav_jobs.append({
                                        "clip_id": clip_id,
                                        "song_url": f"https://suno.com/song/{clip_id}",
                                        "out_wav": wav_expected,
                                        "clip": clip,
                                        "cover_path": clip.get("cover_path"),
                                    })
                                else:
                                    print(f"[DUP] WAV exists, skip queue: {Path(wav_expected).name}")

                        done += 1
                        self.after(0, lambda d=done: self.progress.config(value=d))
                        self.after(0, lambda d=done, t=total: self.progress_label.config(text=f"{d}/{t}"))

                    if do_pl_idx and name not in processed_playlists and not self.stop_event.is_set():
                        xlsx_path = os.path.join(dest, f"{name}.xlsx")
                        id_to_title = {str(c["id"]): str(c["index_title"]) for c in clips if c.get("id")}

                        self._set_status(f"Updating playlist index for {name}…")

                        if os.path.exists(xlsx_path):
                            update_index_titles_in_place(xlsx_path, name, id_to_title)
                            appended = append_missing_rows_preserve_thumbs(xlsx_path, name, clips)
                            print(f"[XLSX] Playlist appended={appended}")
                        else:
                            create_index_xlsx(clips, xlsx_path, name)

                        processed_playlists.add(name)

                    all_clips.extend(clips)

                except Exception as e:
                    print(f"[ERROR] processing input “{line}”: {e}", file=sys.stderr)
                    print(traceback.format_exc(), file=sys.stderr)
                    continue

            if do_master and all_clips and not self.stop_event.is_set():
                master_path = os.path.join(folder, "Suno Master Index.xlsx")
                unique_all = unique_clips_by_id(all_clips)
                id_to_title = {str(c["id"]): str(c["index_title"]) for c in unique_all if c.get("id")}

                self._set_status("Updating master index…")

                if os.path.exists(master_path):
                    update_index_titles_in_place(master_path, "Master", id_to_title)
                    appended = append_missing_rows_preserve_thumbs(master_path, "Master", unique_all)
                    print(f"[XLSX] Master appended={appended}")
                else:
                    create_index_xlsx(unique_all, master_path, "Master")

            if effective_do_wav and not self.stop_event.is_set():
                print(f"[WAV] queued jobs: {len(wav_jobs)}")
                self._set_status(f"WAV: queued {len(wav_jobs)} song(s).")

                if not PLAYWRIGHT_AVAILABLE:
                    self.after(0, lambda: messagebox.showerror(
                        "Missing dependency",
                        "playwright is not installed.\n\nInstall:\n"
                        "  py -m pip install --upgrade playwright\n"
                        "  py -m playwright install"
                    ))
                elif wav_jobs:
                    try:
                        self.download_wavs_in_playwright(wav_jobs, folder)
                        if self.stop_event.is_set():
                            self._set_status("Stopped.")
                        else:
                            self._set_status("WAV: done.")
                    except Exception as e:
                        if not self.stop_event.is_set():
                            self._set_status(f"WAV: error: {e}")
                            print("[WAV] ERROR:", e, file=sys.stderr)
                            print(traceback.format_exc(), file=sys.stderr)
                            self.after(0, lambda err=str(e): messagebox.showerror("WAV download error", err))

            def _finish_ui():
                self.load_btn.config(state=tk.NORMAL)
                self.load_session_btn.config(state=tk.NORMAL)
                self.save_session_btn.config(state=tk.NORMAL)
                self.login_btn.config(state=tk.NORMAL)
                self.download_btn.config(state=tk.NORMAL)
                self._set_running_controls(False)
                if self.stop_event.is_set():
                    self.status_label.config(text="Stopped.")
                    messagebox.showinfo("Stopped", f"Stopped at {done}/{total}.")
                else:
                    self.status_label.config(text="All done!")
                    messagebox.showinfo("Done", f"Completed {done}/{total} items.")

            self.after(0, _finish_ui)

        except Exception as e:
            print("[FATAL] download_worker crashed:", e, file=sys.stderr)
            print(traceback.format_exc(), file=sys.stderr)

            def _fail_ui():
                messagebox.showerror("Fatal error", str(e))
                self.load_btn.config(state=tk.NORMAL)
                self.load_session_btn.config(state=tk.NORMAL)
                self.save_session_btn.config(state=tk.NORMAL)
                self.login_btn.config(state=tk.NORMAL)
                self.download_btn.config(state=tk.NORMAL)
                self._set_running_controls(False)

            try:
                self.after(0, _fail_ui)
            except Exception:
                pass

    def on_close(self):
        try:
            self.stop_event.set()
            self.pause_event.clear()
        except Exception:
            pass
        self.destroy()


if __name__ == "__main__":
    _maybe_run_watcher_mode_from_args()
    setup_profile_logging()
    first_launch_bootstrap_then_exit_if_needed()
    SunoDownloader().mainloop()
